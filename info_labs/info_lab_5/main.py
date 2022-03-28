import openpyxl

file = openpyxl.load_workbook(filename='lab_5.xlsm', data_only=True)  # open
if file is not None:
    print('1234567898765432123456789')
wsh = file['Лист1']  # open work sheet
curr_act = file.active
if curr_act is not None:
    print('1234567898765432123456789')
for x in range(7, 26):  # cycle for change fill color
    for y in range(3, 7):
        if wsh.cell(row=y, column=x).value == 1:
            curr_act.cell(row=y, column=x).fill = openpyxl.styles.PatternFill(start_color='000000', end_color='000000',
                                                                              fill_type='solid')
        if wsh.cell(row=y, column=x).value == 0:
            curr_act.cell(row=y, column=x).fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF',
                                                                              fill_type='solid')
file.close()
file.save('lab_5.xlsm')
# Done by Lagus Maksim
