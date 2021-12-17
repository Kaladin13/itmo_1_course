import openpyxl

file = openpyxl.load_workbook(filename='Lab5.xlsx', data_only=True) #open
wsh = file['Лист1'] #open work sheet
curr_act = file.active
for x in range (7,26): #cycle for change fill color
    for y in range (3,7):
        if wsh.cell(row=y,column=x).value == 0:
            curr_act.cell(row=y,column=x).fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000',fill_type='solid')
        if wsh.cell(row=y,column=x).value == 1:
            curr_act.cell(row=y,column=x).font = openpyxl.styles.Font(italic= True)
file.close()
file.save('Lab5.xlsx')
#By Lagus Maksim